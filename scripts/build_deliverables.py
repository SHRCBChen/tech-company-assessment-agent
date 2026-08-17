from __future__ import annotations

import argparse
import json
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from strict_validation import validate_record


FIELDS = [
    "核心人员", "核心人员背景", "核心人员公开联系方式", "主要产品", "核心技术及应用场景",
    "产业化进展及客户线索", "科创资质及科技项目", "知识产权及技术成果", "上下游",
    "竞争对手", "融资及投资机构背景",
]
RATING = {"A": "优先调研", "B": "重点跟踪", "C": "持续观察", "D": "信息待补"}


def clean(value: object) -> str:
    return re.sub(r"\n{3,}", "\n\n", str(value or "").replace("\r", "")).strip()


def field_text(record: dict, field: str) -> str:
    value = record.get("field_outputs", {}).get(field, {})
    return clean(value.get("text") if isinstance(value, dict) else value)


def sentence(value: str) -> str:
    value = clean(value).replace("｜", "：")
    return value if not value or value.endswith(("。", "！", "？")) else value + "。"


def report(record: dict) -> str:
    assessment = record.get("assessment", {})
    rating = assessment.get("rating", "D")
    sections = [f"# {record['enterprise_name']}企业初评报告", f"**报告日期：** {date.today().isoformat()}　**初步分层：{rating}类（{RATING.get(rating, '信息待补')}）**"]
    groups = [
        ("企业定位与产品逻辑", ["主要产品", "核心技术及应用场景"]),
        ("团队与研发基础", ["核心人员", "核心人员背景", "科创资质及科技项目", "知识产权及技术成果"]),
        ("产业化进展与市场位置", ["产业化进展及客户线索", "上下游", "竞争对手"]),
        ("融资与资本支持", ["融资及投资机构背景"]),
    ]
    numbers = "一二三四五六七八"
    index = 0
    for title, fields in groups:
        facts = [field_text(record, field) for field in fields if field_text(record, field)]
        if facts:
            sections.append(f"## {numbers[index]}、{title}\n\n" + "\n\n".join(sentence(item) for item in facts))
            index += 1
    chains = assessment.get("conclusion_chains", [])
    analysis = []
    for chain in chains:
        text = sentence(chain.get("conclusion", "")) + sentence(chain.get("reasoning", ""))
        conditions = [clean(item) for item in chain.get("change_conditions", []) if clean(item)]
        if conditions:
            text += sentence("可能改变当前判断的关键条件包括" + "；".join(conditions))
        if text:
            analysis.append(text)
    summary = clean(assessment.get("summary"))
    if summary:
        analysis.append(sentence(summary))
    priorities = [clean(item) for item in assessment.get("visit_priorities", []) if clean(item)]
    final = "\n\n".join(analysis) or "当前信息不足以形成可靠评价。"
    final += f"\n\n据此，企业初步分层为{rating}类（{RATING.get(rating, '信息待补')}）。"
    if priorities:
        final += "\n\n建议后续拜访围绕以下问题展开：\n\n" + "\n".join(f"- {item}" for item in priorities)
    sections.append(f"## {numbers[index]}、综合评价与拜访重点\n\n{final}")
    sections.append("本报告仅供初步筛选与拜访准备参考，不构成授信、投资或合作结论。")
    return "\n\n".join(sections) + "\n"


def dossier(record: dict) -> str:
    assessment = record.get("assessment", {})
    rating = assessment.get("rating", "D")
    fact_rows = []
    for fact in record.get("facts", []):
        if fact.get("valid_status") == "已替代":
            continue
        values = [fact.get("fact_id"), fact.get("fact_text"), fact.get("fact_type"), fact.get("source_id"), fact.get("evidence_status"), fact.get("confidence_note")]
        fact_rows.append("| " + " | ".join(clean(value).replace("|", "｜").replace("\n", "；") or "—" for value in values) + " |")
    source_blocks = []
    for source in record.get("sources", []):
        block = [f"### {clean(source.get('source_id'))}｜{clean(source.get('title'))}", f"- 来源：{clean(source.get('source_type'))}", f"- 链接/保存位置：{clean(source.get('location'))}", f"- 支持范围：{clean(source.get('supports'))}"]
        if clean(source.get("cannot_support")):
            block.append(f"- 不能支持：{clean(source.get('cannot_support'))}")
        source_blocks.append("\n\n".join(block))
    onsite = "\n\n".join(clean(item.get("raw_text")) for item in record.get("initial_materials", []) if clean(item.get("raw_text")))
    return (
        f"# {record['enterprise_name']}｜当前事实底稿\n\n"
        f"## 1. 当前画像\n\n- 初步评级：{rating}（{RATING.get(rating, '信息待补')}）\n- 评级理由：{clean(assessment.get('summary'))}\n\n"
        "## 2. 原子事实\n\n| 事实ID | 原子事实 | 类型 | 来源ID | 证据状态 | 边界说明 |\n|---|---|---|---|---|---|\n"
        + ("\n".join(fact_rows) or "| — | — | — | — | — | — |")
        + "\n\n## 3. 证据登记\n\n" + ("\n\n".join(source_blocks) or "暂无来源登记。")
        + (f"\n\n## 4. 大赛现场信息\n\n{onsite}" if onsite else "") + "\n"
    )


def validate(record: dict) -> list[str]:
    issues = []
    assessment = record.get("assessment", {})
    if record.get("status") != "researched":
        issues.append("status is not researched")
    if assessment.get("rating") not in RATING:
        issues.append("invalid rating")
    if not clean(assessment.get("track")) or not clean(assessment.get("development_stage")):
        issues.append("missing track or development_stage")
    chains = assessment.get("conclusion_chains", [])
    if assessment.get("rating") != "D" and len(chains) < 3:
        issues.append("A/B/C report requires at least three conclusion chains")
    for chain in chains:
        if not clean(chain.get("conclusion")) or not clean(chain.get("reasoning")) or not chain.get("basis_fact_ids"):
            issues.append("invalid conclusion chain")
    issues.extend(validate_record(record, FIELDS))
    return list(dict.fromkeys(issues))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run", required=True)
    args = parser.parse_args()
    run_dir = Path(args.run).resolve()
    manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8-sig"))
    records = []
    blocked = {}
    for company in manifest["companies"]:
        record = json.loads((run_dir / company["record_file"]).read_text(encoding="utf-8-sig"))
        issues = validate(record)
        if issues:
            blocked[record["enterprise_name"]] = issues
        records.append(record)
    review_dir = run_dir / "review"
    review_dir.mkdir(parents=True, exist_ok=True)
    batch_issues = []
    if manifest.get("missing_onsite_notes"):
        batch_issues.append("企业名单中仍有未绑定现场笔记的企业")
    if manifest.get("unmatched_note_names"):
        batch_issues.append("现场笔记中仍有未匹配的企业/项目名称")
    if manifest.get("ambiguous_word_paragraphs"):
        batch_issues.append("Word现场笔记仍有歧义段落")
    if manifest.get("pending_ocr_files"):
        batch_issues.append("图片或PDF现场笔记尚未完成OCR/人工读取")
    validation = {
        "status": "PASSED" if not blocked and not batch_issues else "BLOCKED",
        "batch_issues": batch_issues,
        "companies": blocked,
    }
    (review_dir / "validation.json").write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    if blocked or batch_issues:
        raise SystemExit(json.dumps(validation, ensure_ascii=False, indent=2))
    output = run_dir / "deliverables"
    excel_dir, report_dir, dossier_dir = output / "excel", output / "reports", output / "ima-ready"
    for directory in (excel_dir, report_dir, dossier_dir):
        directory.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "企业信息"
    headers = ["企业名称", *FIELDS]
    sheet.append(headers)
    for record in records:
        sheet.append([record["enterprise_name"], *[field_text(record, field) for field in FIELDS]])
        (report_dir / f"{record['enterprise_name']}｜企业初评报告.md").write_text(report(record), encoding="utf-8")
        (dossier_dir / f"{record['enterprise_name']}｜当前事实底稿.md").write_text(dossier(record), encoding="utf-8")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="B4C6E7")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    widths = [26, 24, 34, 26, 32, 36, 40, 36, 40, 30, 28, 44]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "B2"
    excel_path = excel_dir / f"{manifest.get('batch_name', '批次')}｜企业信息主表.xlsx"
    workbook.save(excel_path)
    print(json.dumps({"status": "DELIVERABLES_BUILT", "companies": len(records), "excel": str(excel_path), "reports": str(report_dir), "ima_ready": str(dossier_dir)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
