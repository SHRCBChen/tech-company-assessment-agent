from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from docx import Document


FIELDS = [
    "核心人员", "核心人员背景", "核心人员公开联系方式", "主要产品", "核心技术及应用场景",
    "产业化进展及客户线索", "科创资质及科技项目", "知识产权及技术成果", "上下游",
    "竞争对手", "融资及投资机构背景",
]
NAME_HEADERS = ["企业名称", "企业/项目名称", "项目名称", "公司名称"]
NOTE_HEADERS = ["大赛现场自由笔记", "现场笔记", "现场自由笔记", "现场记录", "自由笔记"]


def text(value: object) -> str:
    return str(value or "").strip()


def load_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [{text(k): text(v) for k, v in row.items()} for row in csv.DictReader(handle)]
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [text(value) for value in values[0]]
    return [{headers[index]: text(value) for index, value in enumerate(row) if index < len(headers)} for row in values[1:]]


def name_aliases(name: str) -> set[str]:
    """Return conservative, reviewable aliases for headings in free-form notes."""
    normalized = match_key(name)
    without_place = re.sub(r"[（(][^（）()]{0,20}[）)]", "", normalized)
    without_legal_suffix = re.sub(r"(股份有限公司|有限责任公司|有限公司|集团有限公司|集团)$", "", without_place)
    without_city_prefix = re.sub(r"^(上海|北京|深圳|广州|杭州|苏州|南京|天津|武汉|成都|西安|厦门|合肥|长沙|重庆)", "", without_legal_suffix)
    aliases = {item for item in (normalized, without_place, without_legal_suffix, without_city_prefix) if len(item) >= 3}
    # A competition note commonly uses a short brand such as “碳派” for “碳派科技（上海）有限公司”.
    short_brand = re.sub(r"(科技|技术|信息|网络|智能|微电子|电子|工业|医学|汽车|船舶|软件)$", "", without_city_prefix)
    if len(short_brand) >= 2:
        aliases.add(short_brand)
    return aliases


def load_word_note_blocks(path: Path, names_by_key: dict[str, str]) -> tuple[list[tuple[str, str, str, str]], list[str], list[str]]:
    """Split a .docx free-note document at enterprise-name headings.

    Return matched blocks as (enterprise, text, location, link_status), plus
    unbound paragraph locations and ambiguous heading locations.  Short aliases
    are accepted only if they uniquely map to one enterprise and remain visible
    in the manifest for later review.
    """
    if path.suffix.lower() != ".docx":
        raise SystemExit("现场笔记Word仅支持 .docx；请将旧版 .doc 另存为 .docx 后重试。")
    alias_map: dict[str, set[str]] = {}
    for enterprise in names_by_key.values():
        for alias in name_aliases(enterprise):
            alias_map.setdefault(alias, set()).add(enterprise)
    unique_aliases = {alias: next(iter(values)) for alias, values in alias_map.items() if len(values) == 1}
    document = Document(path)
    paragraphs = [text(paragraph.text) for paragraph in document.paragraphs]
    blocks: list[tuple[str, str, str, str]] = []
    unbound_locations: list[str] = []
    ambiguous_locations: list[str] = []
    current_name = ""
    current_status = ""
    current_start = 0
    current_lines: list[str] = []

    def flush() -> None:
        nonlocal current_name, current_status, current_start, current_lines
        content = "\n".join(line for line in current_lines if line).strip()
        if current_name and content:
            blocks.append((current_name, content, f"paragraphs {current_start}-{current_start + len(current_lines) - 1}", current_status))
        current_name, current_status, current_start, current_lines = "", "", 0, []

    for index, paragraph in enumerate(paragraphs, start=1):
        if not paragraph:
            if current_name:
                current_lines.append("")
            continue
        normalized = match_key(paragraph)
        candidates: list[tuple[int, str, str]] = []
        for alias, enterprise in unique_aliases.items():
            if alias in normalized:
                candidates.append((len(alias), enterprise, "name_bound" if alias == match_key(enterprise) else "name_bound_alias"))
        if candidates:
            max_length = max(item[0] for item in candidates)
            best = {(enterprise, status) for length, enterprise, status in candidates if length == max_length}
            if len(best) == 1:
                enterprise, status = next(iter(best))
                flush()
                current_name, current_status, current_start, current_lines = enterprise, status, index, [paragraph]
                continue
            ambiguous_locations.append(f"paragraph {index}: {paragraph}")
        if current_name:
            current_lines.append(paragraph)
        else:
            unbound_locations.append(f"paragraph {index}: {paragraph}")
    flush()
    return blocks, unbound_locations, ambiguous_locations


def pick_header(headers: set[str], candidates: list[str]) -> str:
    for candidate in candidates:
        if candidate in headers:
            return candidate
    return ""


def safe_name(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*]+', "-", value).strip(" .-")
    return value[:70] or "batch"


def match_key(value: str) -> str:
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", value)).casefold()


def resolve_note_name(note_name: str, names_by_key: dict[str, str]) -> str:
    """Resolve a separate-note name conservatively; aliases bind only when unique."""
    key = match_key(note_name)
    if key in names_by_key:
        return names_by_key[key]
    candidates = set()
    for enterprise in names_by_key.values():
        aliases = name_aliases(enterprise)
        if key in aliases or (len(key) >= 2 and any(key in alias for alias in aliases)):
            candidates.add(enterprise)
    return next(iter(candidates)) if len(candidates) == 1 else ""


def empty_audit() -> dict:
    round_ids = [
        "R1_subject_mapping", "R2_channel_coverage", "R3_field_deepening",
        "R4_anchor_expansion", "R5_gap_and_conflict", "R6_cross_column_and_output",
    ]
    channels = [
        "official_website_or_account", "government_or_park_attachments", "business_profile",
        "visual_search_cards", "patent_standard_ip", "customer_partner_reverse",
        "financing_investor_primary",
    ]
    return {
        "mode": "full_deep_research",
        "onsite_notes_ingested": False,
        "onsite_notes_decomposed": False,
        "onsite_notes_public_crosschecked": False,
        "rounds": {key: {"status": "pending", "queries_or_paths": [], "new_fact_ids": [], "new_anchors": [], "remaining_tasks": []} for key in round_ids},
        "anchor_queue": [], "weak_source_upgrade_queue": [], "conflict_queue": [],
        "coverage_scan_completed": False, "blank_field_followup_completed": False,
        "company_official_channels_checked": False, "government_attachment_lists_checked": False,
        "business_profile_checked": False, "visual_search_cards_checked": False,
        "industrialization_qualification_ip_deepened": False, "cross_column_scan_completed": False,
        "contact_cleanup_completed": False, "investor_background_deepened": False,
        "channel_checks": {key: {"status": "", "paths": [], "source_ids": [], "notes": ""} for key in channels},
        "field_checks": {}, "person_anchors_found": [], "person_anchors_reviewed": [],
        "confirmed_investors": [], "investor_background_completed_for": [], "investor_checks": {},
        "historical_financing_scan": {"status": "", "paths": [], "source_ids": [], "notes": ""},
    }


def new_record(name: str, note: str, input_path: Path, row_number: object, link_status: str) -> dict:
    enterprise_id = hashlib.sha256(name.encode("utf-8")).hexdigest()[:16]
    source_id = "S-ONSITE-001"
    sources = []
    materials = []
    if note:
        sources.append({
            "source_id": source_id, "source_type": "大赛现场", "title": f"{name}大赛现场自由笔记",
            "location": str(input_path), "retrieved_at": datetime.now().isoformat(timespec="seconds"),
            "supports": "现场记录的团队、产品、客户、经营、融资、里程碑及评委意见", "cannot_support": "",
        })
        materials.append({
            "material_id": "M-ONSITE-001", "material_type": "大赛现场自由笔记", "source_id": source_id,
            "linked_enterprise_name": name, "link_status": link_status, "input_row": row_number,
            "raw_text": note, "processing_status": "pending_extraction", "extracted_fact_ids": [],
        })
    return {
        "enterprise_id": enterprise_id, "enterprise_name": name, "input_name": name,
        "status": "pending_research",
        "subject_mapping": {"status": "ambiguous", "legal_entity": "", "project_name": "", "brand_names": [], "former_names": [], "mapping_basis": ""},
        "coverage": "C", "researched_at": "", "initial_materials": materials, "sources": sources, "facts": [],
        "field_outputs": {field: {"text": "", "basis_fact_ids": []} for field in FIELDS},
        "legacy_inheritance": {"required": False, "completed": True, "candidate_files": [], "matched_sources": [], "reverified_fact_ids": [], "regression_report": "", "unresolved_losses": []},
        "research_audit": empty_audit(),
        "assessment": {"rating": "D", "track": "", "development_stage": "", "stage_benchmark": "", "summary": "", "basis_fact_ids": [], "conclusion_chains": [], "uncertainties": [], "visit_priorities": []},
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--notes", action="append", default=[], help="可重复传入：现场笔记 Excel/CSV、Word(.docx) 或图片；例如 --notes 笔记.docx --notes 照片1.jpg")
    parser.add_argument("--name", required=True)
    parser.add_argument("--output-root", default="runs")
    parser.add_argument("--require-onsite-notes", action="store_true")
    args = parser.parse_args()
    input_path = Path(args.input).resolve()
    rows = load_rows(input_path)
    headers = set(rows[0]) if rows else set()
    name_header = pick_header(headers, NAME_HEADERS)
    note_header = pick_header(headers, NOTE_HEADERS)
    if not name_header:
        raise SystemExit(f"Missing enterprise/project name column. Accepted: {NAME_HEADERS}")
    merged: dict[str, list[str]] = {}
    first_row: dict[str, int] = {}
    note_source: dict[str, tuple[Path, object, str]] = {}
    names_by_key: dict[str, str] = {}
    for index, row in enumerate(rows, start=2):
        name = text(row.get(name_header))
        if not name:
            continue
        key = match_key(name)
        if key in names_by_key and names_by_key[key] != name:
            raise SystemExit(f"企业名单存在规范化后重名，无法安全匹配：{names_by_key[key]} / {name}")
        names_by_key[key] = name
        merged.setdefault(name, [])
        note = text(row.get(note_header)) if note_header else ""
        if note:
            merged[name].append(note)
            note_source.setdefault(name, (input_path, index, "row_bound"))
        first_row.setdefault(name, index)

    unmatched_note_names: list[str] = []
    unbound_word_paragraphs: list[str] = []
    ambiguous_word_paragraphs: list[str] = []
    pending_ocr_files: list[str] = []
    notes_paths = [Path(value).resolve() for value in args.notes]
    image_suffixes = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff", ".heic"}
    for notes_path in notes_paths:
        if not notes_path.exists():
            raise SystemExit(f"现场笔记文件不存在：{notes_path}")
        suffix = notes_path.suffix.lower()
        if suffix in {".xlsx", ".csv"}:
            note_rows = load_rows(notes_path)
            note_headers = set(note_rows[0]) if note_rows else set()
            note_name_header = pick_header(note_headers, NAME_HEADERS)
            separate_note_header = pick_header(note_headers, NOTE_HEADERS)
            if not note_name_header or not separate_note_header:
                raise SystemExit(f"现场笔记表格须同时有名称列{NAME_HEADERS}和笔记列{NOTE_HEADERS}：{notes_path.name}")
            seen_unmatched: set[str] = set()
            for index, row in enumerate(note_rows, start=2):
                note_name = text(row.get(note_name_header))
                note = text(row.get(separate_note_header))
                if not note_name or not note:
                    continue
                matched_name = resolve_note_name(note_name, names_by_key)
                if not matched_name:
                    if note_name not in seen_unmatched:
                        unmatched_note_names.append(note_name)
                        seen_unmatched.add(note_name)
                    continue
                if note not in merged[matched_name]:
                    merged[matched_name].append(note)
                note_source.setdefault(matched_name, (notes_path, index, "name_bound"))
        elif suffix == ".docx":
            blocks, unbound, ambiguous = load_word_note_blocks(notes_path, names_by_key)
            unbound_word_paragraphs.extend(f"{notes_path.name} {item}" for item in unbound)
            ambiguous_word_paragraphs.extend(f"{notes_path.name} {item}" for item in ambiguous)
            for matched_name, note, location, link_status in blocks:
                if note not in merged[matched_name]:
                    merged[matched_name].append(note)
                note_source.setdefault(matched_name, (notes_path, location, link_status))
        elif suffix in image_suffixes or suffix == ".pdf":
            # OCR/visual reading is intentionally done by the agent, not guessed from a filename.
            pending_ocr_files.append(str(notes_path))
        else:
            raise SystemExit(f"不支持的现场笔记格式：{notes_path.name}。支持 .xlsx/.csv/.docx 和常见图片/PDF。")

    if args.require_onsite_notes and not note_header and not notes_paths:
        raise SystemExit(f"未提供现场笔记列或独立现场笔记文件。Accepted note headers: {NOTE_HEADERS}")
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    run_dir = Path(args.output_root).resolve() / f"{stamp}-{safe_name(args.name)}"
    records_dir = run_dir / "records"
    records_dir.mkdir(parents=True, exist_ok=False)
    companies = []
    missing_notes = []
    for name, notes in merged.items():
        note = "\n".join(dict.fromkeys(notes))
        if not note:
            missing_notes.append(name)
        source_path, source_row, link_status = note_source.get(name, (input_path, first_row[name], "row_bound"))
        record = new_record(name, note, source_path, source_row, link_status)
        record_file = f"records/{record['enterprise_id']}.json"
        (run_dir / record_file).write_text(json.dumps(record, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        companies.append({"enterprise_id": record["enterprise_id"], "enterprise_name": name, "record_file": record_file})
    manifest = {
        "batch_id": run_dir.name, "batch_name": args.name, "created_at": datetime.now().isoformat(timespec="seconds"),
        "input_file": str(input_path), "notes_files": [str(path) for path in notes_paths],
        "input_binding_mode": "separate_files_by_name" if notes_paths else "combined_file_same_row",
        "company_count": len(companies), "missing_onsite_notes": missing_notes,
        "unmatched_note_names": unmatched_note_names,
        "unbound_word_paragraphs": unbound_word_paragraphs,
        "ambiguous_word_paragraphs": ambiguous_word_paragraphs,
        "pending_ocr_files": pending_ocr_files,
        "companies": companies,
    }
    (run_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    latest = Path(args.output_root).resolve() / "latest-run.txt"
    latest.parent.mkdir(parents=True, exist_ok=True)
    latest.write_text(str(run_dir), encoding="utf-8")
    print(json.dumps({"status": "BATCH_CREATED", "run": str(run_dir), "companies": len(companies), "missing_onsite_notes": missing_notes, "unmatched_note_names": unmatched_note_names, "unbound_word_paragraphs": unbound_word_paragraphs, "ambiguous_word_paragraphs": ambiguous_word_paragraphs, "pending_ocr_files": pending_ocr_files}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
